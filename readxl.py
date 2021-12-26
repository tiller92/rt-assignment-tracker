# central python dict will cont all RT's and then keep points based of where
# they are at in the excel sheet. Use boolean for there core teams, floors,
# and other. To get the assignemnt the charge could paste the file location
# intp an inpit and then click submit. On submit use python to read the exel
# and then place the points accoringly. Decide with team wether it should be
# monthly or per schedule period.
from functions import proccess_cell
import openpyxl 

# Get the RT all staff list
# for dev this is the test list. You will have export funtionality from this file

def excel_break(file):
    """this functon shoule take the assignment and break down who is here and where they where assigned. File will come from a folder that the charge places the current assignment list in. This function will return a dict of who was where that will be used to updated the DB"""
    # path = f"/Users/ryantiller/Desktop/rt-assignment-project/current-assignment/test.xlsx"
    # unblcok this to use website
    path = file

    #To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    
    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.

    # Cell object is created by using
    # sheet object's cell() method.
    cell_obj = sheet_obj.cell(row = 2, column = 2)

    # Print value of cell object
    # using the value attribute
    # print(cell_obj.value)
    # print('*****')
    # print(sheet_obj.max_column, 'columns')
    # print('*****')
    # print(sheet_obj.max_row, 'rows')
    # print('******')
    
    Charge = [sheet_obj.cell(row=2, column=5)]
    Chare_staff = []
    
    MICU = [sheet_obj.cell(row= 3, column=5),sheet_obj.cell(row= 4, column=5),sheet_obj.cell(row= 5, column=5)]
    MICU_staff = []
    
    JMICU = [sheet_obj.cell(row= 7, column=5),sheet_obj.cell(row= 8, column=5),sheet_obj.cell(row= 9, column=5),sheet_obj.cell(row= 10, column=5)]
        #includes JSICU
    JMICU_staff = []
    

    NCCU = [sheet_obj.cell(row=11, column=5),sheet_obj.cell(row= 12, column=5)]
    NCCU_staff = []
    
    TEN_MICU = [sheet_obj.cell(row= 13, column=5),sheet_obj.cell(row= 14, column=5),sheet_obj.cell(row= 15, column=5)]
    TEN_MICU_staff = []
    
    ROSS = [sheet_obj.cell(row= 16, column=5),sheet_obj.cell(row= 17, column=5),sheet_obj.cell(row= 18, column=5),sheet_obj.cell(row= 19, column=5),sheet_obj.cell(row= 20, column=5),sheet_obj.cell(row= 21, column=5),sheet_obj.cell(row= 22, column=5),sheet_obj.cell(row= 23, column=5)]
    ROSS_staff = []
    
    NICU = [sheet_obj.cell(row=2, column=12),sheet_obj.cell(row= 3, column=12)]
    NICU_staff = []
    
    UHERT = [sheet_obj.cell(row=5, column=12)]
    UHERT_staff = []
    
    JERT = [sheet_obj.cell(row=6, column=12)]
    JERT_staff = []
    
    ED = [sheet_obj.cell(row=8, column=12),sheet_obj.cell(row=9, column=12)]
    ED_staff = []
    

    # will include PCU     
    FLOORS = [sheet_obj.cell(row= 11, column=12),sheet_obj.cell(row= 12, column=12),sheet_obj.cell(row= 15, column=12),sheet_obj.cell(row= 16, column=12),sheet_obj.cell(row= 17, column=12),sheet_obj.cell(row= 18, column=12),sheet_obj.cell(row= 19, column=12),sheet_obj.cell(row= 20, column=12),sheet_obj.cell(row= 21, column=12),sheet_obj.cell(row= 22, column=12),]
    FLOORS_staff = []
 
    #make a filtered list that is just strings

        

    for cell in MICU:
        try:
            if cell.value and cell.value != None:
                MICU_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
        
      
    for cell in JMICU:
        try:
            if cell.value and cell.value != None:
                JMICU_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
    
    for cell in NCCU:
        try:
            if cell.value and cell.value != None:
                NCCU_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
    for cell in TEN_MICU:
        try:
            if cell.value and cell.value != None:
                TEN_MICU_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
    
    for cell in ROSS:
        try:
            if cell.value and cell.value != None:
                ROSS_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
    
    for cell in NICU:
        try:
            if cell.value and cell.value != None:
                NICU_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
    
    for cell in UHERT:
        try:
            if cell.value and cell.value != None:
                UHERT_staff.append(cell.value)
   
        except Exception:
                print(Exception)   
                 
    for cell in JERT:
        try:
            if cell.value and cell.value != None:
                JERT_staff.append(cell.value)
   
        except Exception:
                print(Exception)         
        
    for cell in ED:
        try:
            if cell.value and cell.value != None:
                ED_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
        
    for cell in FLOORS:
        try:
            if cell.value and cell.value != None:
                FLOORS_staff.append(cell.value)
   
        except Exception:
                print(Exception) 
                
    for cell in Charge:
        try:
            if cell.value and cell.value != None:
                Chare_staff.append(cell.value)
   
        except Exception:
                print(Exception)


    staffing = {'units':{
        'Charge':Chare_staff,
          'MICU':MICU_staff,
         'JMICU':JMICU_staff,
         '10ICU':TEN_MICU_staff,
         'NCCU':NCCU_staff,
         'ROSS':ROSS_staff,
         'NICU':NICU_staff,
         'ED':ED_staff,
         'UHERT': UHERT_staff,
         'JERT': JERT_staff,
         'FLOORS':FLOORS_staff,

     }}


    return staffing
    